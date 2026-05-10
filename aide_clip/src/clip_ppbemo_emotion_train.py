import argparse
import hashlib
import json
import math
import os
import random
import time
from collections import Counter
from pathlib import Path
from typing import Dict, List, Optional

from openpyxl import load_workbook
from PIL import Image


EMOTION_LABELS = ["AD", "SAD", "FD", "DD", "SD", "HD", "ND"]
EMOTION_DISPLAY_MAP = {
    "AD": "angry driving",
    "SAD": "sad driving",
    "FD": "fear driving",
    "DD": "disgust driving",
    "SD": "surprise driving",
    "HD": "happy driving",
    "ND": "neutral driving",
}
EMOTION_PROMPT_MAP = {
    "AD": "angry",
    "SAD": "sad",
    "FD": "fearful",
    "DD": "disgusted",
    "SD": "surprised",
    "HD": "happy",
    "ND": "neutral",
}

VISUAL_COLUMN_MAP = {
    "face_crgb": "PPB_Emo_dataset@video-face-CRGB",
    "face_lrgb": "PPB_Emo_dataset@video-face-LRGB",
    "face_rrgb": "PPB_Emo_dataset@video-face-RRGB",
    "face_cir": "PPB_Emo_dataset@video-face-CIR",
    "body": "PPB_Emo_dataset@video-body",
    "road": "PPB_Emo_dataset@video-road",
}

PROJECT_ROOT = Path(__file__).resolve().parent.parent
DEFAULT_DATA_ROOT = os.environ.get("PPBEMO_ROOT", str(PROJECT_ROOT / "data" / "bbpemo"))
DEFAULT_OUTPUT = str(PROJECT_ROOT / "results" / "ppbemo" / "clip_ppbemo_emotion_supervised_results.json")


def resolve_default_annotation_xlsx(data_root: str) -> str:
    candidates = [
        Path(data_root) / "Psychological_data" / "Emotion_label.xlsx",
        Path(data_root) / "Psychological_data" / "Emotion label.xlsx",
    ]
    for candidate in candidates:
        if candidate.exists():
            return str(candidate)
    return str(candidates[0])


DEFAULT_ANNOTATION_XLSX = resolve_default_annotation_xlsx(DEFAULT_DATA_ROOT)


def normalize_emotion_label(label: str) -> Optional[str]:
    key = str(label).strip().upper()
    return key if key in EMOTION_LABELS else None


def log(message: str) -> None:
    print(f"[{time.strftime('%Y-%m-%d %H:%M:%S')}] {message}", flush=True)


def format_duration(seconds: float) -> str:
    total = max(0, int(seconds))
    hours, rem = divmod(total, 3600)
    minutes, secs = divmod(rem, 60)
    if hours > 0:
        return f"{hours:02d}:{minutes:02d}:{secs:02d}"
    return f"{minutes:02d}:{secs:02d}"


def default_checkpoint_path(output_path: Path) -> Path:
    return output_path.with_suffix(".ckpt.pt")


def clone_module_state_dict(module) -> Dict[str, object]:
    return {key: value.detach().cpu().clone() for key, value in module.state_dict().items()}


def accuracy(y_true: List[str], y_pred: List[str]) -> float:
    if not y_true:
        return 0.0
    return sum(1 for t, p in zip(y_true, y_pred) if t == p) / len(y_true)


def weighted_f1(y_true: List[str], y_pred: List[str], labels: List[str]) -> float:
    if not y_true:
        return 0.0
    support = Counter(y_true)
    total = len(y_true)
    weighted = 0.0
    for label in labels:
        tp = sum(1 for t, p in zip(y_true, y_pred) if t == label and p == label)
        fp = sum(1 for t, p in zip(y_true, y_pred) if t != label and p == label)
        fn = sum(1 for t, p in zip(y_true, y_pred) if t == label and p != label)
        precision = tp / (tp + fp) if (tp + fp) > 0 else 0.0
        recall = tp / (tp + fn) if (tp + fn) > 0 else 0.0
        f1 = 2 * precision * recall / (precision + recall) if (precision + recall) > 0 else 0.0
        weighted += (support.get(label, 0) / total) * f1
    return weighted


def confusion_matrix(y_true: List[str], y_pred: List[str], labels: List[str]) -> Dict[str, Dict[str, int]]:
    mat = {t: {p: 0 for p in labels} for t in labels}
    for t, p in zip(y_true, y_pred):
        if t in mat and p in mat[t]:
            mat[t][p] += 1
    return mat


def evaluate_split(y_true: List[str], y_pred: List[str]) -> Dict:
    return {
        "accuracy": round(accuracy(y_true, y_pred), 6),
        "weighted_f1": round(weighted_f1(y_true, y_pred, EMOTION_LABELS), 6),
        "confusion_matrix": confusion_matrix(y_true, y_pred, EMOTION_LABELS),
    }


def unique_group_ids(samples: List[Dict], group_key: str) -> List[str]:
    return sorted({str(sample[group_key]) for sample in samples})


def normalize_header_name(value) -> str:
    return str(value or "").replace("\n", " ").strip()


def index_video_files(data_root: str) -> Dict[str, str]:
    root = Path(data_root)
    if not root.exists():
        raise FileNotFoundError(f"PPB-Emo root not found: {root}")

    mapping: Dict[str, str] = {}
    for path in sorted(root.rglob("*.mp4")):
        mapping.setdefault(path.stem, str(path))
        mapping.setdefault(path.name, str(path))
    return mapping


def load_annotation_rows(annotation_xlsx: str) -> List[Dict[str, object]]:
    if not os.path.exists(annotation_xlsx):
        raise FileNotFoundError(f"Annotation file not found: {annotation_xlsx}")

    workbook = load_workbook(annotation_xlsx, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = sheet.iter_rows(values_only=True)
    try:
        header = [normalize_header_name(x) for x in next(rows)]
    except StopIteration as exc:
        raise RuntimeError(f"Empty annotation workbook: {annotation_xlsx}") from exc

    records = []
    for row_idx, row in enumerate(rows, start=2):
        item = {header[col_idx]: row[col_idx] for col_idx in range(min(len(header), len(row)))}
        item["__row_index__"] = row_idx
        records.append(item)
    return records


def collect_samples(
    data_root: str,
    annotation_xlsx: str,
    video_column: str,
    max_sequences: int = 0,
) -> List[Dict]:
    if video_column not in VISUAL_COLUMN_MAP:
        raise ValueError(f"Unsupported video_column: {video_column}. Choices: {sorted(VISUAL_COLUMN_MAP)}")

    records = load_annotation_rows(annotation_xlsx)
    video_index = index_video_files(data_root)
    excel_video_column = VISUAL_COLUMN_MAP[video_column]
    candidate_set = set(EMOTION_LABELS)

    samples: List[Dict] = []
    missing_video = 0
    missing_label = 0

    for record in records:
        label = normalize_emotion_label(record.get("category", ""))
        if label not in candidate_set:
            missing_label += 1
            continue

        video_key = str(record.get(excel_video_column, "") or "").strip()
        if not video_key:
            missing_video += 1
            continue

        video_path = video_index.get(video_key) or video_index.get(f"{video_key}.mp4")
        if not video_path:
            missing_video += 1
            continue

        sample = {
            "sequence_id": f"row{record['__row_index__']}_{record.get('participant', 'unknown')}_{label}",
            "row_index": int(record["__row_index__"]),
            "participant": str(record.get("participant", "")).strip(),
            "valence": record.get("valence"),
            "arousal": record.get("arousal"),
            "dominance": record.get("dominance"),
            "intensity": record.get("intensity"),
            "label": label,
            "label_text": EMOTION_DISPLAY_MAP[label],
            "video_column": video_column,
            "video_key": video_key,
            "video_path": video_path,
            "modalities": {
                normalize_header_name(key): value
                for key, value in record.items()
                if isinstance(key, str) and key.startswith("PPB_Emo_dataset@")
            },
        }
        samples.append(sample)

    log(
        f"[DATA] rows={len(records)} valid_samples={len(samples)} "
        f"missing_label={missing_label} missing_video={missing_video} video_column={video_column}"
    )

    if max_sequences > 0:
        samples = samples[:max_sequences]
    return samples


def split_samples(
    samples: List[Dict],
    train_ratio: float,
    val_ratio: float,
    seed: int,
    split_mode: str,
) -> Dict[str, List[Dict]]:
    if split_mode == "participant":
        participant_groups: Dict[str, List[Dict]] = {}
        for sample in samples:
            participant_groups.setdefault(str(sample["participant"]), []).append(sample)

        rng = random.Random(seed)
        participant_ids = list(participant_groups.keys())
        rng.shuffle(participant_ids)

        total_samples = len(samples)
        target_train = total_samples * train_ratio
        target_val = total_samples * val_ratio
        train, val, test = [], [], []
        train_count = 0
        val_count = 0

        for participant_id in participant_ids:
            group = list(participant_groups[participant_id])
            if train_count < target_train:
                train.extend(group)
                train_count += len(group)
            elif val_count < target_val:
                val.extend(group)
                val_count += len(group)
            else:
                test.extend(group)

        rng.shuffle(train)
        rng.shuffle(val)
        rng.shuffle(test)
        return {"train": train, "val": val, "test": test}

    label_groups: Dict[str, List[Dict]] = {}
    for sample in samples:
        label_groups.setdefault(sample["label"], []).append(sample)

    rng = random.Random(seed)
    train, val, test = [], [], []
    for _, group in label_groups.items():
        group = list(group)
        rng.shuffle(group)
        n = len(group)
        n_train = int(n * train_ratio)
        n_val = int(n * val_ratio)
        n_test = n - n_train - n_val
        train.extend(group[:n_train])
        val.extend(group[n_train:n_train + n_val])
        test.extend(group[n_train + n_val:n_train + n_val + n_test])

    rng.shuffle(train)
    rng.shuffle(val)
    rng.shuffle(test)
    return {"train": train, "val": val, "test": test}


def build_prompt_templates(prompt_template: str, prompt_set: str) -> List[str]:
    if prompt_set == "single":
        return [prompt_template]
    if prompt_set == "default_5":
        return [
            "The driver looks <LABEL>.",
            "The driver appears <LABEL>.",
            "The person behind the wheel seems <LABEL>.",
            "This in-car video shows a <LABEL> driver.",
            "The driver's facial expression looks <LABEL>.",
        ]
    if prompt_set == "ppbemo_natural_7":
        return [
            "The driver looks <LABEL>.",
            "The driver appears <LABEL>.",
            "The driver seems <LABEL> while driving.",
            "The person behind the wheel looks <LABEL>.",
            "This in-car video shows a driver who looks <LABEL>.",
            "The driver's facial expression appears <LABEL>.",
            "The visible emotion in this driving clip looks <LABEL>.",
        ]
    custom = [x.strip() for x in prompt_set.split("||") if x.strip()]
    return custom if custom else [prompt_template]


def build_class_prompts(prompt_template: str, prompt_set: str) -> List[List[str]]:
    templates = build_prompt_templates(prompt_template, prompt_set)
    return [[tpl.replace("<LABEL>", EMOTION_PROMPT_MAP[label]) for tpl in templates] for label in EMOTION_LABELS]


def build_prompt_group_indices(num_prompts: int, group_size: int) -> List[List[int]]:
    if group_size <= 0 or group_size >= num_prompts:
        return [list(range(num_prompts))]
    groups = []
    for start in range(0, num_prompts, group_size):
        groups.append(list(range(start, min(start + group_size, num_prompts))))
    return groups


def sample_frame_indices(total_frames: int, num_frames: int) -> List[int]:
    if total_frames <= 0:
        return [0]
    if num_frames <= 1:
        return [max(0, total_frames // 2)]
    if total_frames <= num_frames:
        return list(range(total_frames))
    return [round(index * (total_frames - 1) / (num_frames - 1)) for index in range(num_frames)]


def read_sampled_frames(video_path: str, num_frames: int) -> List[Image.Image]:
    import cv2

    capture = cv2.VideoCapture(video_path)
    if not capture.isOpened():
        raise RuntimeError(f"Failed to open video: {video_path}")

    total_frames = int(capture.get(cv2.CAP_PROP_FRAME_COUNT))
    target_indices = sample_frame_indices(total_frames, num_frames)
    images: List[Image.Image] = []

    for frame_index in target_indices:
        capture.set(cv2.CAP_PROP_POS_FRAMES, frame_index)
        ok, frame = capture.read()
        if not ok or frame is None:
            continue
        rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        images.append(Image.fromarray(rgb))

    if not images:
        capture.set(cv2.CAP_PROP_POS_FRAMES, 0)
        ok, frame = capture.read()
        if ok and frame is not None:
            rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
            images.append(Image.fromarray(rgb))

    capture.release()

    if not images:
        raise RuntimeError(f"Failed to decode frames from video: {video_path}")
    if num_frames > 0 and len(images) < num_frames:
        last = images[-1]
        while len(images) < num_frames:
            images.append(last.copy())
    return images


def read_middle_frame(video_path: str) -> Image.Image:
    return read_sampled_frames(video_path, 1)[0]


def extract_image_features(
    samples: List[Dict],
    processor,
    model,
    device: str,
    batch_size: int,
    num_frames: int,
    tag: str,
    return_frame_features: bool = False,
):
    import torch

    feats = []
    total_batches = max(1, math.ceil(len(samples) / batch_size))
    start_time = time.time()
    log(f"[FEATURES] start {tag}: samples={len(samples)}, batches={total_batches}, num_frames={num_frames}")
    for batch_idx, start in enumerate(range(0, len(samples), batch_size), start=1):
        batch = samples[start:start + batch_size]
        batch_features = []
        for sample in batch:
            images = read_sampled_frames(sample["video_path"], num_frames)
            inputs = processor(images=images, return_tensors="pt", padding=True).to(device)
            with torch.no_grad():
                image_features = model.get_image_features(pixel_values=inputs["pixel_values"])
                image_features = image_features / image_features.norm(dim=-1, keepdim=True).clamp(min=1e-12)
            if return_frame_features:
                batch_features.append(image_features.float().cpu())
            else:
                pooled = image_features.mean(dim=0)
                pooled = pooled / pooled.norm(dim=-1, keepdim=True).clamp(min=1e-12)
                batch_features.append(pooled)
        feats.append(torch.stack(batch_features, dim=0).float().cpu())
        if batch_idx == 1 or batch_idx % 10 == 0 or batch_idx == total_batches:
            elapsed = time.time() - start_time
            eta = elapsed / batch_idx * (total_batches - batch_idx)
            log(
                f"[FEATURES] {tag}: batch {batch_idx}/{total_batches}, "
                f"elapsed={format_duration(elapsed)}, eta={format_duration(eta)}"
            )
    return torch.cat(feats, dim=0)


def extract_text_features(prompt_groups: List[List[str]], processor, model, device: str):
    import torch

    class_prompt_features = []
    log(
        f"[TEXT] start text feature extraction: classes={len(prompt_groups)}, "
        f"prompts_per_class={len(prompt_groups[0]) if prompt_groups else 0}"
    )
    for prompts in prompt_groups:
        inputs = processor(text=prompts, return_tensors="pt", padding=True).to(device)
        with torch.no_grad():
            text_features = model.get_text_features(input_ids=inputs["input_ids"], attention_mask=inputs["attention_mask"])
            text_features = text_features / text_features.norm(dim=-1, keepdim=True).clamp(min=1e-12)
        class_prompt_features.append(text_features)
    log("[TEXT] done text feature extraction")
    return torch.stack(class_prompt_features, dim=0).float().detach()


class ClipImageAdapter:
    def __init__(
        self,
        dim: int,
        device: str,
        hidden_dim: int,
        dropout: float,
        num_classes: int,
        num_prompts: int,
        use_prompt_weight: bool = True,
        use_class_temperature: bool = True,
        use_class_bias: bool = True,
    ):
        import torch
        import torch.nn as nn

        self.device = device
        self.input_proj = nn.Linear(dim, hidden_dim).to(device)
        self.net = nn.Sequential(
            nn.LayerNorm(hidden_dim),
            nn.Linear(hidden_dim, hidden_dim),
            nn.GELU(),
            nn.Dropout(dropout),
            nn.Linear(hidden_dim, hidden_dim),
        ).to(device)
        self.out_proj = nn.Linear(hidden_dim, dim).to(device)
        nn.init.zeros_(self.out_proj.weight)
        nn.init.zeros_(self.out_proj.bias)

        self.logit_scale = nn.Parameter(torch.tensor(1.0, device=device))
        self.prompt_weight_logits = nn.Parameter(torch.zeros(num_classes, num_prompts, device=device))
        self.class_logit_scale = nn.Parameter(torch.zeros(num_classes, device=device))
        self.class_bias = nn.Parameter(torch.zeros(num_classes, device=device))
        self.use_prompt_weight = use_prompt_weight
        self.use_class_temperature = use_class_temperature
        self.use_class_bias = use_class_bias

    def parameters(self):
        return (
            list(self.input_proj.parameters())
            + list(self.net.parameters())
            + list(self.out_proj.parameters())
            + [self.logit_scale, self.prompt_weight_logits, self.class_logit_scale, self.class_bias]
        )

    def state_dict(self):
        return {
            "input_proj": clone_module_state_dict(self.input_proj),
            "net": clone_module_state_dict(self.net),
            "out_proj": clone_module_state_dict(self.out_proj),
            "logit_scale": self.logit_scale.detach().cpu().clone(),
            "prompt_weight_logits": self.prompt_weight_logits.detach().cpu().clone(),
            "class_logit_scale": self.class_logit_scale.detach().cpu().clone(),
            "class_bias": self.class_bias.detach().cpu().clone(),
            "use_prompt_weight": self.use_prompt_weight,
            "use_class_temperature": self.use_class_temperature,
            "use_class_bias": self.use_class_bias,
        }

    def load_state_dict(self, state):
        self.input_proj.load_state_dict(state["input_proj"])
        self.net.load_state_dict(state["net"])
        self.out_proj.load_state_dict(state["out_proj"])
        self.logit_scale.data.copy_(state["logit_scale"].to(self.device))
        self.prompt_weight_logits.data.copy_(state["prompt_weight_logits"].to(self.device))
        self.class_logit_scale.data.copy_(state["class_logit_scale"].to(self.device))
        self.class_bias.data.copy_(state["class_bias"].to(self.device))
        self.use_prompt_weight = state.get("use_prompt_weight", True)
        self.use_class_temperature = state.get("use_class_temperature", True)
        self.use_class_bias = state.get("use_class_bias", True)

    def train(self):
        self.input_proj.train()
        self.net.train()
        self.out_proj.train()

    def eval(self):
        self.input_proj.eval()
        self.net.eval()
        self.out_proj.eval()

    def _adapt_image(self, image_x):
        base = self.input_proj(image_x)
        delta = self.net(base)
        fused = base + delta
        img = image_x + self.out_proj(fused)
        return img / img.norm(dim=-1, keepdim=True).clamp(min=1e-12)

    def logits(self, image_x, text_x):
        import torch
        import torch.nn.functional as F

        img = self._adapt_image(image_x)
        txt = text_x / text_x.norm(dim=-1, keepdim=True).clamp(min=1e-12)
        sim = torch.einsum("bd,cpd->bcp", img, txt)

        if self.use_prompt_weight:
            prompt_w = F.softmax(self.prompt_weight_logits, dim=-1).unsqueeze(0)
            class_sim = (sim * prompt_w).sum(dim=-1)
        else:
            class_sim = sim.mean(dim=-1)

        global_scale = self.logit_scale.exp().clamp(max=100.0)
        if self.use_class_temperature:
            class_scale = self.class_logit_scale.exp().clamp(min=0.5, max=2.5).unsqueeze(0)
        else:
            class_scale = 1.0
        if self.use_class_bias:
            class_bias = self.class_bias.unsqueeze(0)
        else:
            class_bias = 0.0
        return global_scale * class_sim * class_scale + class_bias

    def grouped_logits(self, image_x, text_x, group_indices: List[List[int]]):
        import torch

        img = self._adapt_image(image_x)
        txt = text_x / text_x.norm(dim=-1, keepdim=True).clamp(min=1e-12)
        sim = torch.einsum("bd,cpd->bcp", img, txt)

        group_scores = []
        for gidx in group_indices:
            group_scores.append(sim[:, :, gidx].mean(dim=-1))
        scores = torch.stack(group_scores, dim=-1)

        global_scale = self.logit_scale.exp().clamp(max=100.0)
        if self.use_class_temperature:
            class_scale = self.class_logit_scale.exp().clamp(min=0.5, max=2.5).view(1, -1, 1)
        else:
            class_scale = 1.0
        if self.use_class_bias:
            class_bias = self.class_bias.view(1, -1, 1)
        else:
            class_bias = 0.0
        return global_scale * scores * class_scale + class_bias


def predict_emotion_from_features(
    image_features,
    text_features,
    adapter,
    idx2label: Dict[int, str],
    batch_size: int,
    use_test_ensemble: bool,
    ensemble_group_size: int,
) -> List[str]:
    import torch

    preds = []
    group_indices = build_prompt_group_indices(int(text_features.shape[1]), ensemble_group_size)

    for start in range(0, image_features.shape[0], batch_size):
        batch_x = image_features[start:start + batch_size]
        with torch.no_grad():
            if use_test_ensemble and len(group_indices) > 1:
                g_logits = adapter.grouped_logits(batch_x.to(adapter.device), text_features.to(adapter.device), group_indices)
                group_pred = g_logits.argmax(dim=1)
                total_scores = g_logits.sum(dim=-1)
                idxs = []
                for i in range(group_pred.shape[0]):
                    votes = torch.bincount(group_pred[i], minlength=len(idx2label))
                    top = votes.max()
                    cands = (votes == top).nonzero(as_tuple=False).view(-1)
                    if cands.numel() == 1:
                        idxs.append(int(cands[0].item()))
                    else:
                        cs = total_scores[i, cands]
                        idxs.append(int(cands[cs.argmax()].item()))
            else:
                logits = adapter.logits(batch_x.to(adapter.device), text_features.to(adapter.device))
                idxs = logits.argmax(dim=-1).detach().cpu().tolist()
        preds.extend([idx2label[i] for i in idxs])
    return preds


def predict_zeroshot_from_features(
    image_features,
    text_features,
    idx2label: Dict[int, str],
    batch_size: int,
    use_test_ensemble: bool,
    ensemble_group_size: int,
) -> List[str]:
    import torch

    preds = []
    group_indices = build_prompt_group_indices(int(text_features.shape[1]), ensemble_group_size)

    for start in range(0, image_features.shape[0], batch_size):
        batch_x = image_features[start:start + batch_size]
        device = batch_x.device
        img = batch_x / batch_x.norm(dim=-1, keepdim=True).clamp(min=1e-12)
        text_x = text_features.to(device)
        text_x = text_x / text_x.norm(dim=-1, keepdim=True).clamp(min=1e-12)
        sim = torch.einsum("bd,cpd->bcp", img, text_x)
        if use_test_ensemble and len(group_indices) > 1:
            grouped = torch.stack([sim[:, :, gidx].mean(dim=-1) for gidx in group_indices], dim=-1)
            group_pred = grouped.argmax(dim=1)
            total_scores = grouped.sum(dim=-1)
            idxs = []
            for i in range(group_pred.shape[0]):
                votes = torch.bincount(group_pred[i], minlength=len(idx2label))
                top = votes.max()
                cands = (votes == top).nonzero(as_tuple=False).view(-1)
                if cands.numel() == 1:
                    idxs.append(int(cands[0].item()))
                else:
                    cs = total_scores[i, cands]
                    idxs.append(int(cands[cs.argmax()].item()))
        else:
            idxs = sim.mean(dim=-1).argmax(dim=-1).tolist()
        preds.extend([idx2label[i] for i in idxs])
    return preds


def predict_zeroshot_from_frame_features(
    frame_features,
    text_features,
    idx2label: Dict[int, str],
    batch_size: int,
    frame_aggregation: str,
) -> List[str]:
    import torch

    preds = []
    for start in range(0, frame_features.shape[0], batch_size):
        batch_x = frame_features[start:start + batch_size]
        device = batch_x.device
        img = batch_x / batch_x.norm(dim=-1, keepdim=True).clamp(min=1e-12)
        text_x = text_features.to(device)
        text_x = text_x / text_x.norm(dim=-1, keepdim=True).clamp(min=1e-12)
        sim = torch.einsum("bfd,cpd->bfcp", img, text_x)
        class_frame_scores = sim.mean(dim=-1)
        if frame_aggregation == "max":
            video_scores = class_frame_scores.max(dim=1).values
        else:
            video_scores = class_frame_scores.mean(dim=1)
        idxs = video_scores.argmax(dim=-1).detach().cpu().tolist()
        preds.extend([idx2label[i] for i in idxs])
    return preds


def predict_emotion(samples: List[Dict], processor, model, prompts: List[str], device: str, batch_size: int) -> List[str]:
    import torch

    preds = []
    for start in range(0, len(samples), batch_size):
        batch = samples[start:start + batch_size]
        images = [read_middle_frame(sample["video_path"]) for sample in batch]
        inputs = processor(text=prompts, images=images, return_tensors="pt", padding=True).to(device)
        with torch.no_grad():
            outputs = model(**inputs)
            logits = outputs.logits_per_image
            idxs = logits.argmax(dim=-1).detach().cpu().tolist()
        preds.extend([EMOTION_LABELS[i] for i in idxs])
    return preds


def train_clip_supervised(
    model,
    processor,
    train_samples: List[Dict],
    val_samples: List[Dict],
    prompts: List[str],
    device: str,
    epochs: int,
    batch_size: int,
    lr: float,
    weight_decay: float,
    max_grad_norm: float,
):
    import torch
    import torch.nn as nn

    label2idx = {label: idx for idx, label in enumerate(EMOTION_LABELS)}
    optimizer = torch.optim.AdamW(model.parameters(), lr=lr, weight_decay=weight_decay)
    criterion = nn.CrossEntropyLoss()

    best_state = None
    best_val_acc = -1.0
    overall_start = time.time()

    for epoch_idx in range(epochs):
        model.train()
        random.shuffle(train_samples)
        epoch_start = time.time()
        running_loss = 0.0
        num_batches = 0

        for start in range(0, len(train_samples), batch_size):
            batch = train_samples[start:start + batch_size]
            images = [read_middle_frame(sample["video_path"]) for sample in batch]
            targets = torch.tensor([label2idx[sample["label"]] for sample in batch], dtype=torch.long, device=device)
            inputs = processor(text=prompts, images=images, return_tensors="pt", padding=True).to(device)

            outputs = model(**inputs)
            logits = outputs.logits_per_image
            loss = criterion(logits, targets)

            optimizer.zero_grad()
            loss.backward()
            if max_grad_norm > 0:
                torch.nn.utils.clip_grad_norm_(model.parameters(), max_grad_norm)
            optimizer.step()
            running_loss += float(loss.item())
            num_batches += 1

        model.eval()
        with torch.no_grad():
            val_pred = predict_emotion(val_samples, processor, model, prompts, device, batch_size)
            val_true = [sample["label"] for sample in val_samples]
            val_acc = accuracy(val_true, val_pred)

        if val_acc > best_val_acc:
            best_val_acc = val_acc
            best_state = {k: v.detach().cpu().clone() for k, v in model.state_dict().items()}

        elapsed = time.time() - overall_start
        eta = elapsed / (epoch_idx + 1) * (epochs - epoch_idx - 1)
        log(
            f"[TRAIN] epoch {epoch_idx + 1}/{epochs} | loss={running_loss / max(1, num_batches):.6f} | "
            f"val_acc={val_acc:.6f} | best_val_acc={best_val_acc:.6f} | "
            f"epoch_time={format_duration(time.time() - epoch_start)} | elapsed={format_duration(elapsed)} | eta={format_duration(eta)}"
        )

    if best_state is not None:
        model.load_state_dict(best_state)
    model.eval()
    return model


def train_strict_frozen_clip(
    train_x,
    train_y,
    val_x,
    val_y,
    text_features,
    adapter: ClipImageAdapter,
    epochs: int,
    batch_size: int,
    lr: float,
    weight_decay: float,
    max_grad_norm: float,
    use_class_weight: bool,
    label_smoothing: float,
    select_metric: str,
    use_test_ensemble: bool,
    ensemble_group_size: int,
):
    import torch
    import torch.nn as nn

    class_weights = None
    if use_class_weight:
        class_counts = torch.bincount(train_y, minlength=len(EMOTION_LABELS)).float()
        class_weights = (class_counts.sum() / class_counts.clamp(min=1.0)).to(adapter.device)
        class_weights = class_weights / class_weights.mean().clamp(min=1e-12)

    criterion = nn.CrossEntropyLoss(weight=class_weights, label_smoothing=label_smoothing)
    optimizer = torch.optim.AdamW(adapter.parameters(), lr=lr, weight_decay=weight_decay)

    idx2label = {idx: label for idx, label in enumerate(EMOTION_LABELS)}
    overall_start = time.time()

    adapter.eval()
    with torch.no_grad():
        init_val_pred = predict_emotion_from_features(
            val_x,
            text_features,
            adapter,
            idx2label,
            batch_size,
            use_test_ensemble=use_test_ensemble,
            ensemble_group_size=ensemble_group_size,
        )
        init_val_true = [EMOTION_LABELS[int(x.item())] for x in val_y]
        init_val_acc = accuracy(init_val_true, init_val_pred)
        init_val_wf1 = weighted_f1(init_val_true, init_val_pred, EMOTION_LABELS)
        best_val_metric = init_val_wf1 if select_metric == "weighted_f1" else init_val_acc
    best_state = adapter.state_dict()
    log(
        f"[TRAIN] epoch 0/{epochs} | loss=nan | val_acc={init_val_acc:.6f} | "
        f"val_wf1={init_val_wf1:.6f} | best_metric={best_val_metric:.6f} | "
        f"epoch_time=00:00 | elapsed=00:00 | eta=--:--"
    )

    for epoch_idx in range(epochs):
        adapter.train()
        epoch_start = time.time()
        running_loss = 0.0
        num_batches = 0
        perm = torch.randperm(train_x.shape[0])
        train_x = train_x[perm]
        train_y = train_y[perm]

        for start in range(0, train_x.shape[0], batch_size):
            bx = train_x[start:start + batch_size].to(adapter.device)
            by = train_y[start:start + batch_size].to(adapter.device)
            logits = adapter.logits(bx, text_features.to(adapter.device))
            loss = criterion(logits, by)

            optimizer.zero_grad()
            loss.backward()
            if max_grad_norm > 0:
                torch.nn.utils.clip_grad_norm_(adapter.parameters(), max_grad_norm)
            optimizer.step()
            running_loss += float(loss.item())
            num_batches += 1

        adapter.eval()
        with torch.no_grad():
            val_pred = predict_emotion_from_features(
                val_x,
                text_features,
                adapter,
                idx2label,
                batch_size,
                use_test_ensemble=use_test_ensemble,
                ensemble_group_size=ensemble_group_size,
            )
            val_true = [EMOTION_LABELS[int(x.item())] for x in val_y]
            val_acc = accuracy(val_true, val_pred)
            val_wf1 = weighted_f1(val_true, val_pred, EMOTION_LABELS)
            metric = val_wf1 if select_metric == "weighted_f1" else val_acc

        if metric > best_val_metric:
            best_val_metric = metric
            best_state = adapter.state_dict()

        elapsed = time.time() - overall_start
        eta = elapsed / (epoch_idx + 1) * (epochs - epoch_idx - 1)
        log(
            f"[TRAIN] epoch {epoch_idx + 1}/{epochs} | loss={running_loss / max(1, num_batches):.6f} | "
            f"val_acc={val_acc:.6f} | val_wf1={val_wf1:.6f} | best_metric={best_val_metric:.6f} | "
            f"epoch_time={format_duration(time.time() - epoch_start)} | elapsed={format_duration(elapsed)} | eta={format_duration(eta)}"
        )

    if best_state is not None:
        adapter.load_state_dict(best_state)
    adapter.eval()
    return adapter


def train_linear_probe(
    train_x,
    train_y,
    val_x,
    val_y,
    epochs: int,
    batch_size: int,
    lr: float,
    weight_decay: float,
    max_grad_norm: float,
    use_class_weight: bool,
    label_smoothing: float,
    select_metric: str,
    device: str,
):
    import torch
    import torch.nn as nn

    classifier = nn.Linear(int(train_x.shape[1]), len(EMOTION_LABELS)).to(device)

    class_weights = None
    if use_class_weight:
        class_counts = torch.bincount(train_y, minlength=len(EMOTION_LABELS)).float()
        class_weights = (class_counts.sum() / class_counts.clamp(min=1.0)).to(device)
        class_weights = class_weights / class_weights.mean().clamp(min=1e-12)

    criterion = nn.CrossEntropyLoss(weight=class_weights, label_smoothing=label_smoothing)
    optimizer = torch.optim.AdamW(classifier.parameters(), lr=lr, weight_decay=weight_decay)

    idx2label = {idx: label for idx, label in enumerate(EMOTION_LABELS)}
    best_state = None
    best_val_metric = -1.0
    overall_start = time.time()

    def predict_logits(features):
        preds = []
        for start in range(0, features.shape[0], batch_size):
            bx = features[start:start + batch_size].to(device)
            with torch.no_grad():
                logits = classifier(bx)
                idxs = logits.argmax(dim=-1).detach().cpu().tolist()
            preds.extend([idx2label[i] for i in idxs])
        return preds

    classifier.eval()
    with torch.no_grad():
        init_val_pred = predict_logits(val_x)
        init_val_true = [EMOTION_LABELS[int(x.item())] for x in val_y]
        init_val_acc = accuracy(init_val_true, init_val_pred)
        init_val_wf1 = weighted_f1(init_val_true, init_val_pred, EMOTION_LABELS)
        best_val_metric = init_val_wf1 if select_metric == "weighted_f1" else init_val_acc
    best_state = {k: v.detach().cpu().clone() for k, v in classifier.state_dict().items()}
    log(
        f"[LINEAR] epoch 0/{epochs} | loss=nan | val_acc={init_val_acc:.6f} | "
        f"val_wf1={init_val_wf1:.6f} | best_metric={best_val_metric:.6f} | "
        f"epoch_time=00:00 | elapsed=00:00 | eta=--:--"
    )

    for epoch_idx in range(epochs):
        classifier.train()
        epoch_start = time.time()
        running_loss = 0.0
        num_batches = 0
        perm = torch.randperm(train_x.shape[0])
        train_x = train_x[perm]
        train_y = train_y[perm]

        for start in range(0, train_x.shape[0], batch_size):
            bx = train_x[start:start + batch_size].to(device)
            by = train_y[start:start + batch_size].to(device)
            logits = classifier(bx)
            loss = criterion(logits, by)

            optimizer.zero_grad()
            loss.backward()
            if max_grad_norm > 0:
                torch.nn.utils.clip_grad_norm_(classifier.parameters(), max_grad_norm)
            optimizer.step()
            running_loss += float(loss.item())
            num_batches += 1

        classifier.eval()
        with torch.no_grad():
            val_pred = predict_logits(val_x)
            val_true = [EMOTION_LABELS[int(x.item())] for x in val_y]
            val_acc = accuracy(val_true, val_pred)
            val_wf1 = weighted_f1(val_true, val_pred, EMOTION_LABELS)
            metric = val_wf1 if select_metric == "weighted_f1" else val_acc

        if metric > best_val_metric:
            best_val_metric = metric
            best_state = {k: v.detach().cpu().clone() for k, v in classifier.state_dict().items()}

        elapsed = time.time() - overall_start
        eta = elapsed / (epoch_idx + 1) * (epochs - epoch_idx - 1)
        log(
            f"[LINEAR] epoch {epoch_idx + 1}/{epochs} | loss={running_loss / max(1, num_batches):.6f} | "
            f"val_acc={val_acc:.6f} | val_wf1={val_wf1:.6f} | best_metric={best_val_metric:.6f} | "
            f"epoch_time={format_duration(time.time() - epoch_start)} | elapsed={format_duration(elapsed)} | eta={format_duration(eta)}"
        )

    if best_state is not None:
        classifier.load_state_dict(best_state)
    classifier.eval()
    return classifier


def parse_args():
    parser = argparse.ArgumentParser(description="Standalone PPB-Emo CLIP emotion training")
    parser.set_defaults(strict_frozen_clip=True)
    parser.add_argument("--data_root", default=DEFAULT_DATA_ROOT)
    parser.add_argument("--annotation_xlsx", default=DEFAULT_ANNOTATION_XLSX)
    parser.add_argument("--video_column", choices=sorted(VISUAL_COLUMN_MAP.keys()), default="face_crgb")
    parser.add_argument("--split_mode", choices=["participant", "sample_stratified"], default="participant")
    parser.add_argument("--train_ratio", type=float, default=0.65)
    parser.add_argument("--val_ratio", type=float, default=0.15)
    parser.add_argument("--seed", type=int, default=42)
    parser.add_argument("--max_sequences", type=int, default=0)
    parser.add_argument("--device", default="cuda:0")
    parser.add_argument("--clip_mode", choices=["offline_only", "auto"], default="auto")
    parser.add_argument("--model_id", default="openai/clip-vit-base-patch32")
    parser.add_argument("--prompt_template", default="The driver looks <LABEL>.")
    parser.add_argument("--prompt_set", default="ppbemo_natural_7", help="single | default_5 | ppbemo_natural_7 | custom templates joined by ||")
    parser.add_argument("--epochs", type=int, default=40)
    parser.add_argument("--batch_size", type=int, default=32)
    parser.add_argument("--lr", type=float, default=1.5e-4)
    parser.add_argument("--weight_decay", type=float, default=5e-4)
    parser.add_argument("--max_grad_norm", type=float, default=1.0)
    parser.add_argument("--num_frames", type=int, default=3)
    parser.add_argument("--adapter_hidden_dim", type=int, default=1024)
    parser.add_argument("--adapter_dropout", type=float, default=0.2)
    parser.add_argument("--use_class_weight", action="store_true")
    parser.add_argument("--label_smoothing", type=float, default=0.05)
    parser.add_argument("--select_metric", choices=["accuracy", "weighted_f1"], default="weighted_f1")
    parser.add_argument("--use_test_ensemble", action="store_true")
    parser.add_argument("--ensemble_group_size", type=int, default=2)
    parser.add_argument("--strict_frozen_clip", dest="strict_frozen_clip", action="store_true", help="Use frozen CLIP + adapter pipeline (default)")
    parser.add_argument("--full_finetune_clip", dest="strict_frozen_clip", action="store_false", help="Disable frozen CLIP adapter pipeline and finetune CLIP directly")
    parser.add_argument("--strict_strategy", choices=["adapter", "linear_probe", "zeroshot"], default="adapter")
    parser.add_argument("--video_aggregation", choices=["mean_pool", "frame_mean", "frame_max"], default="mean_pool")
    parser.add_argument("--disable_prompt_weight", action="store_true")
    parser.add_argument("--disable_class_temperature", action="store_true")
    parser.add_argument("--disable_class_bias", action="store_true")
    parser.add_argument("--feature_cache_dir", default=None)
    parser.add_argument("--checkpoint_output", default=None)
    parser.add_argument("--output", default=DEFAULT_OUTPUT)
    return parser.parse_args()


def main():
    args = parse_args()
    random.seed(args.seed)

    samples = collect_samples(args.data_root, args.annotation_xlsx, args.video_column, args.max_sequences)
    if len(samples) < 10:
        raise RuntimeError(f"Too few valid samples: {len(samples)}")
    log(f"[INFO] valid samples: {len(samples)}")

    splits = split_samples(samples, args.train_ratio, args.val_ratio, args.seed, args.split_mode)
    train_samples = splits["train"]
    val_samples = splits["val"]
    test_samples = splits["test"]
    log(f"[INFO] split sizes -> train: {len(train_samples)}, val: {len(val_samples)}, test: {len(test_samples)}")
    log(
        f"[SPLIT] mode={args.split_mode} "
        f"train_participants={unique_group_ids(train_samples, 'participant')} "
        f"val_participants={unique_group_ids(val_samples, 'participant')} "
        f"test_participants={unique_group_ids(test_samples, 'participant')}"
    )

    prompt_groups = build_class_prompts(args.prompt_template, args.prompt_set)
    single_prompts = [x[0] for x in prompt_groups]
    prompts_per_class = len(prompt_groups[0]) if prompt_groups else 0
    total_text_prompts = sum(len(group) for group in prompt_groups)
    if args.strict_frozen_clip:
        pipeline_name = f"strict_frozen_clip_{args.strict_strategy}"
    else:
        pipeline_name = "full_clip_finetune"
    log(
        f"[PIPELINE] mode={pipeline_name} num_frames={args.num_frames} "
        f"prompts_per_class={prompts_per_class} total_text_prompts={total_text_prompts}"
    )
    log(
        f"[PROMPT] classes={len(prompt_groups)} prompts_per_class={len(prompt_groups[0]) if prompt_groups else 0} "
        f"total_prompts={len(prompt_groups) * (len(prompt_groups[0]) if prompt_groups else 0)}"
    )

    import torch
    from transformers import CLIPModel, CLIPProcessor

    if args.clip_mode == "auto":
        try:
            processor = CLIPProcessor.from_pretrained(args.model_id)
            model = CLIPModel.from_pretrained(args.model_id, use_safetensors=False)
        except Exception:
            processor = CLIPProcessor.from_pretrained(args.model_id, local_files_only=True)
            model = CLIPModel.from_pretrained(args.model_id, use_safetensors=False, local_files_only=True)
    else:
        processor = CLIPProcessor.from_pretrained(args.model_id, local_files_only=True)
        model = CLIPModel.from_pretrained(args.model_id, use_safetensors=False, local_files_only=True)

    dtype = torch.float16 if args.device.startswith("cuda") else torch.float32
    model = model.to(device=args.device, dtype=dtype)
    log(f"[INFO] model loaded: {args.model_id} on {args.device}")

    label2idx = {label: idx for idx, label in enumerate(EMOTION_LABELS)}
    idx2label = {idx: label for idx, label in enumerate(EMOTION_LABELS)}

    if args.strict_frozen_clip:
        model.eval()
        for p in model.parameters():
            p.requires_grad = False

        cache_path = None
        if args.feature_cache_dir:
            os.makedirs(args.feature_cache_dir, exist_ok=True)
            cache_key_src = {
                "data_root": os.path.abspath(args.data_root),
                "annotation_xlsx": os.path.abspath(args.annotation_xlsx),
                "video_column": args.video_column,
                "split_mode": args.split_mode,
                "train_ratio": args.train_ratio,
                "val_ratio": args.val_ratio,
                "seed": args.seed,
                "max_sequences": args.max_sequences,
                "model_id": args.model_id,
                "prompt_template": args.prompt_template,
                "prompt_set": args.prompt_set,
                "num_frames": args.num_frames,
                "video_aggregation": args.video_aggregation,
            }
            cache_key = hashlib.sha1(json.dumps(cache_key_src, sort_keys=True).encode("utf-8")).hexdigest()[:16]
            cache_path = os.path.join(args.feature_cache_dir, f"strict_features_{cache_key}.pt")

        if cache_path and os.path.exists(cache_path):
            log(f"[CACHE] loading strict features from: {cache_path}")
            cached = torch.load(cache_path, map_location="cpu")
            train_x = cached["train_x"]
            val_x = cached["val_x"]
            test_x = cached["test_x"]
            train_frame_x = cached.get("train_frame_x")
            val_frame_x = cached.get("val_frame_x")
            test_frame_x = cached.get("test_frame_x")
            text_features = cached["text_features"]
            train_y = cached["train_y"]
            val_y = cached["val_y"]
            test_y = cached["test_y"]
        else:
            log("[CACHE] miss" if cache_path else "[CACHE] disabled")
            return_frame_features = args.video_aggregation in {"frame_mean", "frame_max"}
            train_features = extract_image_features(
                train_samples, processor, model, args.device, args.batch_size, args.num_frames, tag="train", return_frame_features=return_frame_features
            )
            val_features = extract_image_features(
                val_samples, processor, model, args.device, args.batch_size, args.num_frames, tag="val", return_frame_features=return_frame_features
            )
            test_features = extract_image_features(
                test_samples, processor, model, args.device, args.batch_size, args.num_frames, tag="test", return_frame_features=return_frame_features
            )
            if return_frame_features:
                train_frame_x = train_features
                val_frame_x = val_features
                test_frame_x = test_features
                train_x = train_frame_x.mean(dim=1)
                train_x = train_x / train_x.norm(dim=-1, keepdim=True).clamp(min=1e-12)
                val_x = val_frame_x.mean(dim=1)
                val_x = val_x / val_x.norm(dim=-1, keepdim=True).clamp(min=1e-12)
                test_x = test_frame_x.mean(dim=1)
                test_x = test_x / test_x.norm(dim=-1, keepdim=True).clamp(min=1e-12)
            else:
                train_x = train_features
                val_x = val_features
                test_x = test_features
                train_frame_x = None
                val_frame_x = None
                test_frame_x = None
            text_features = extract_text_features(prompt_groups, processor, model, args.device)

            train_y = torch.tensor([label2idx[s["label"]] for s in train_samples], dtype=torch.long)
            val_y = torch.tensor([label2idx[s["label"]] for s in val_samples], dtype=torch.long)
            test_y = torch.tensor([label2idx[s["label"]] for s in test_samples], dtype=torch.long)

            if cache_path:
                torch.save(
                    {
                        "train_x": train_x,
                        "val_x": val_x,
                        "test_x": test_x,
                        "train_frame_x": train_frame_x,
                        "val_frame_x": val_frame_x,
                        "test_frame_x": test_frame_x,
                        "text_features": text_features,
                        "train_y": train_y,
                        "val_y": val_y,
                        "test_y": test_y,
                    },
                    cache_path,
                )
                log(f"[CACHE] saved strict features to: {cache_path}")

        if args.video_aggregation == "frame_mean":
            zero_shot_val_pred = predict_zeroshot_from_frame_features(
                val_frame_x,
                text_features,
                idx2label,
                args.batch_size,
                frame_aggregation="mean",
            )
            zero_shot_test_pred = predict_zeroshot_from_frame_features(
                test_frame_x,
                text_features,
                idx2label,
                args.batch_size,
                frame_aggregation="mean",
            )
        elif args.video_aggregation == "frame_max":
            zero_shot_val_pred = predict_zeroshot_from_frame_features(
                val_frame_x,
                text_features,
                idx2label,
                args.batch_size,
                frame_aggregation="max",
            )
            zero_shot_test_pred = predict_zeroshot_from_frame_features(
                test_frame_x,
                text_features,
                idx2label,
                args.batch_size,
                frame_aggregation="max",
            )
        else:
            zero_shot_val_pred = predict_zeroshot_from_features(
                val_x,
                text_features,
                idx2label,
                args.batch_size,
                use_test_ensemble=args.use_test_ensemble,
                ensemble_group_size=args.ensemble_group_size,
            )
            zero_shot_test_pred = predict_zeroshot_from_features(
                test_x,
                text_features,
                idx2label,
                args.batch_size,
                use_test_ensemble=args.use_test_ensemble,
                ensemble_group_size=args.ensemble_group_size,
            )
        adapter = None
        linear_probe = None
        if args.strict_strategy == "adapter":
            adapter = ClipImageAdapter(
                dim=int(train_x.shape[1]),
                device=args.device,
                hidden_dim=args.adapter_hidden_dim,
                dropout=args.adapter_dropout,
                num_classes=len(EMOTION_LABELS),
                num_prompts=int(text_features.shape[1]),
                use_prompt_weight=not args.disable_prompt_weight,
                use_class_temperature=not args.disable_class_temperature,
                use_class_bias=not args.disable_class_bias,
            )
            adapter = train_strict_frozen_clip(
                train_x=train_x,
                train_y=train_y,
                val_x=val_x,
                val_y=val_y,
                text_features=text_features,
                adapter=adapter,
                epochs=args.epochs,
                batch_size=args.batch_size,
                lr=args.lr,
                weight_decay=args.weight_decay,
                max_grad_norm=args.max_grad_norm,
                use_class_weight=args.use_class_weight,
                label_smoothing=args.label_smoothing,
                select_metric=args.select_metric,
                use_test_ensemble=args.use_test_ensemble,
                ensemble_group_size=args.ensemble_group_size,
            )
            val_pred = predict_emotion_from_features(
                val_x,
                text_features,
                adapter,
                idx2label,
                args.batch_size,
                use_test_ensemble=args.use_test_ensemble,
                ensemble_group_size=args.ensemble_group_size,
            )
            test_pred = predict_emotion_from_features(
                test_x,
                text_features,
                adapter,
                idx2label,
                args.batch_size,
                use_test_ensemble=args.use_test_ensemble,
                ensemble_group_size=args.ensemble_group_size,
            )
        elif args.strict_strategy == "linear_probe":
            linear_probe = train_linear_probe(
                train_x=train_x,
                train_y=train_y,
                val_x=val_x,
                val_y=val_y,
                epochs=args.epochs,
                batch_size=args.batch_size,
                lr=args.lr,
                weight_decay=args.weight_decay,
                max_grad_norm=args.max_grad_norm,
                use_class_weight=args.use_class_weight,
                label_smoothing=args.label_smoothing,
                select_metric=args.select_metric,
                device=args.device,
            )

            def predict_linear(features):
                preds = []
                for start in range(0, features.shape[0], args.batch_size):
                    bx = features[start:start + args.batch_size].to(args.device)
                    with torch.no_grad():
                        idxs = linear_probe(bx).argmax(dim=-1).detach().cpu().tolist()
                    preds.extend([idx2label[i] for i in idxs])
                return preds

            val_pred = predict_linear(val_x)
            test_pred = predict_linear(test_x)
        else:
            val_pred = zero_shot_val_pred
            test_pred = zero_shot_test_pred

        val_true = [EMOTION_LABELS[int(i.item())] for i in val_y]
        test_true = [EMOTION_LABELS[int(i.item())] for i in test_y]
        zero_shot_result = {
            "val": evaluate_split(val_true, zero_shot_val_pred),
            "test": evaluate_split(test_true, zero_shot_test_pred),
        }
    else:
        model = train_clip_supervised(
            model=model,
            processor=processor,
            train_samples=train_samples,
            val_samples=val_samples,
            prompts=single_prompts,
            device=args.device,
            epochs=args.epochs,
            batch_size=args.batch_size,
            lr=args.lr,
            weight_decay=args.weight_decay,
            max_grad_norm=args.max_grad_norm,
        )
        val_pred = predict_emotion(val_samples, processor, model, single_prompts, args.device, args.batch_size)
        test_pred = predict_emotion(test_samples, processor, model, single_prompts, args.device, args.batch_size)
        val_true = [s["label"] for s in val_samples]
        test_true = [s["label"] for s in test_samples]
        zero_shot_result = None

    checkpoint_path = Path(args.checkpoint_output) if args.checkpoint_output else default_checkpoint_path(Path(args.output))

    result = {
        "config": {
            "method": "clip_supervised_text_image_emotion",
            "task": "emotion",
            "dataset": "PPB-Emo",
            "annotation_xlsx": args.annotation_xlsx,
            "video_column": args.video_column,
            "split_mode": args.split_mode,
            "num_classes": len(EMOTION_LABELS),
            "split": {
                "train": args.train_ratio,
                "val": args.val_ratio,
                "test": round(1 - args.train_ratio - args.val_ratio, 6),
            },
            "model_id": args.model_id,
            "prompt_template": args.prompt_template,
            "prompt_set": args.prompt_set,
            "prompts_per_class": prompts_per_class,
            "total_text_prompts": total_text_prompts,
            "pipeline_mode": pipeline_name,
            "epochs": args.epochs,
            "batch_size": args.batch_size,
            "lr": args.lr,
            "weight_decay": args.weight_decay,
            "max_grad_norm": args.max_grad_norm,
            "num_frames": args.num_frames,
            "adapter_hidden_dim": args.adapter_hidden_dim,
            "adapter_dropout": args.adapter_dropout,
            "use_class_weight": args.use_class_weight,
            "label_smoothing": args.label_smoothing,
            "select_metric": args.select_metric,
            "use_test_ensemble": args.use_test_ensemble,
            "ensemble_group_size": args.ensemble_group_size,
            "strict_frozen_clip": args.strict_frozen_clip,
            "strict_strategy": args.strict_strategy,
            "video_aggregation": args.video_aggregation,
            "use_prompt_weight": not args.disable_prompt_weight,
            "use_class_temperature": not args.disable_class_temperature,
            "use_class_bias": not args.disable_class_bias,
            "feature_cache_dir": args.feature_cache_dir,
            "checkpoint_output": str(checkpoint_path),
            "seed": args.seed,
            "max_sequences": args.max_sequences,
            "prompt_groups_count": len(prompt_groups),
            "prompts_per_class": len(prompt_groups[0]) if prompt_groups else 0,
            "total_prompt_texts": len(prompt_groups) * (len(prompt_groups[0]) if prompt_groups else 0),
        },
        "dataset": {
            "total": len(samples),
            "train": len(train_samples),
            "val": len(val_samples),
            "test": len(test_samples),
            "label_distribution_total": dict(Counter([s["label"] for s in samples])),
            "label_distribution_train": dict(Counter([s["label"] for s in train_samples])),
            "participants_train": unique_group_ids(train_samples, "participant"),
            "participants_val": unique_group_ids(val_samples, "participant"),
            "participants_test": unique_group_ids(test_samples, "participant"),
        },
        "label_map": EMOTION_DISPLAY_MAP,
        "prompt_groups": prompt_groups,
        "val": evaluate_split(val_true, val_pred),
        "test": evaluate_split(test_true, test_pred),
    }
    if zero_shot_result is not None:
        result["zero_shot"] = zero_shot_result

    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False, indent=2)

    checkpoint_path.parent.mkdir(parents=True, exist_ok=True)
    checkpoint_payload = {
        "config": result["config"],
        "dataset": result["dataset"],
        "metrics": {
            "val": result["val"],
            "test": result["test"],
        },
        "prompt_groups": prompt_groups,
        "label2idx": label2idx,
        "idx2label": idx2label,
        "output_path": str(output_path),
        "label_map": EMOTION_DISPLAY_MAP,
    }
    if args.strict_frozen_clip and args.strict_strategy == "adapter":
        checkpoint_payload.update(
            {
                "checkpoint_type": "strict_frozen_clip_adapter",
                "adapter_state_dict": adapter.state_dict(),
                "text_features": text_features.cpu(),
            }
        )
    elif args.strict_frozen_clip and args.strict_strategy == "linear_probe":
        checkpoint_payload.update(
            {
                "checkpoint_type": "strict_frozen_clip_linear_probe",
                "linear_probe_state_dict": {k: v.detach().cpu().clone() for k, v in linear_probe.state_dict().items()},
                "text_features": text_features.cpu(),
            }
        )
    elif args.strict_frozen_clip:
        checkpoint_payload.update(
            {
                "checkpoint_type": "strict_frozen_clip_zeroshot",
                "text_features": text_features.cpu(),
            }
        )
    else:
        checkpoint_payload.update(
            {
                "checkpoint_type": "clip_finetune_model",
                "model_state_dict": {k: v.detach().cpu().clone() for k, v in model.state_dict().items()},
            }
        )
    torch.save(checkpoint_payload, checkpoint_path)

    log(f"[DONE] saved supervised CLIP emotion report to: {output_path}")
    log(f"[DONE] saved checkpoint to: {checkpoint_path}")
    print(json.dumps({"test": result["test"]}, ensure_ascii=False, indent=2), flush=True)


if __name__ == "__main__":
    main()